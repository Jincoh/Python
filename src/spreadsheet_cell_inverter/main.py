from pathlib import Path

import openpyxl
from openpyxl.utils import cell, column_index_from_string

def main():
    wb = openpyxl.load_workbook("wb.xlsx")
    sheet = wb.active
    if sheet == None:
        return

    riter = sheet.iter_rows()

    lst = []
    idx = 0
    for row in riter:
        lst.append([])
        for col in row:
            lst[idx].append(col.value)
        idx+=1
    olist = []
    for row in range(len(lst)):
        for col in range(len(lst[row])):
            olist.append([])
            for x in range(len(lst)):
                olist[col].append(lst[x][col])
            if col >= len(lst[row])-1:
                break
        else:
            continue
        break

    sheet.delete_rows(1,sheet.max_row)
    for row in range(len(olist)):
        for col in range(len(olist[row])):

            sheet.cell(row=row+1, column=col+1).value = olist[row][col]
    
    wb.save("owb.xlsx")

if __name__ == "__main__":
    main()
