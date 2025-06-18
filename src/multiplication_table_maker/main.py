import openpyxl

from pathlib import Path
from openpyxl.styles import Font
from openpyxl.utils.cell import get_column_letter

def main():
    path = Path("mtable.xlsx")

    wb = openpyxl.load_workbook(path)
    font = Font(bold=True)
    
    sheet = wb.active
    if sheet == None:
        return

    for i in range(1,11):
        sheet.cell(column=1, row  = i+1).value = i
        sheet.cell(column = i+1, row = 1).value = i
    for i in range(2,12):
        for j in range(2,12):
            letter = get_column_letter(j)
            sheet.cell(row=i, column = j).value=f"={letter}1 * A{i}"

    wb.save("mtable2.xlsx")

if __name__ == "__main__":
    main()
